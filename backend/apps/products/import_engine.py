"""
Streaming, chunked Excel/CSV product import engine.

Runs on a background thread kicked off by the upload view (see import_views.py).
Never loads a whole workbook into memory: .xlsx/.csv are read row-by-row via
openpyxl's read_only mode / the stdlib csv module, accumulated into fixed-size
chunks, and each chunk is validated + bulk-written in its own transaction so a
single bad batch never aborts the rest of the import.
"""
import csv
import datetime
import time
import uuid

from django.db import transaction, close_old_connections, connections
from openpyxl import load_workbook

from .models import Product, ImportBatch, ImportErrorLog
from .views import infer_category_name

CHUNK_SIZE = 500

# Same fuzzy column-name matching the previous synchronous importer used,
# just resolved once per file (against the header row) instead of once per
# row per field, which is what makes this viable at 100k+ row scale.
FIELD_ALIASES = {
    'sku': ['code', 'sku', 'item code'],
    'model_no': ['modelno', 'model_no', 'model'],
    'name': ['item_name', 'product name', 'product', 'name', 'particular'],
    'name_local': ['item_name_l', 'local name', 'name local'],
    'barcode': ['barcode', 'upc', 'tag'],
    'category': ['category name', 'category'],
    'sub_category': ['sub category', 'sub_category', 'subcat'],
    'group': ['group', 'group_name'],
    'brand': ['brand', 'make', 'company'],
    'color_code': ['colorcode', 'color_code', 'color'],
    'size': ['size'],
    'supplier': ['supplier', 'vendor'],
    'address': ['address', 'location'],
    'cost_price': ['cost price', 'cost', 'cp', 'purchase price'],
    'selling_price': ['sales price', 'selling price', 'sp', 'mrp'],
    'stock': ['stock qty', 'stock', 'qty'],
}

BLANK_TOKENS = ('', 'nan', 'none', 'null', 'undefined', 'n/a', '-')


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        s = str(val).strip().replace('$', '').replace('₹', '').replace(',', '')
        if not s or s.lower() in ('none', 'null', 'nan', 'undefined', 'n/a', '-', 'free'):
            return default
        return float(s)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=1):
    if val is None:
        return default
    try:
        s = str(val).strip().replace(',', '')
        if not s or s.lower() in ('none', 'null', 'nan', 'undefined', 'n/a', '-', 'out of stock'):
            return default
        return int(float(s))
    except (ValueError, TypeError):
        return default


def detect_ext(filename):
    lower = (filename or '').lower()
    if lower.endswith('.csv'):
        return 'csv'
    if lower.endswith('.xlsx'):
        return 'xlsx'
    if lower.endswith('.xls'):
        return 'xls'
    return None


def resolve_field_headers(headers):
    """
    Map each logical field to the one real header it matches, computed once per file.
    Two global passes (not one pass per field): every field's *exact* alias match is
    resolved and the header claimed before any field is allowed to fall back to a
    substring match. Otherwise a generic alias like "code" (for sku) matches inside
    "Barcode" via substring before "SKU/Code" is ever considered, since "Barcode" sorts
    earlier — silently mapping sku and barcode to the same column.
    """
    lower_headers = [(h, h.lower().strip()) for h in headers]
    resolved = {}
    claimed = set()

    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            alias_l = alias.lower().strip()
            match = next((h for h, hl in lower_headers if hl == alias_l and h not in claimed), None)
            if match:
                resolved[field] = match
                claimed.add(match)
                break

    for field, aliases in FIELD_ALIASES.items():
        if field in resolved:
            continue
        for alias in aliases:
            alias_l = alias.lower().strip()
            match = next((h for h, hl in lower_headers if alias_l in hl and h not in claimed), None)
            if match:
                resolved[field] = match
                claimed.add(match)
                break

    return resolved


def field_getter(resolved_headers, row_dict):
    def get(field):
        h = resolved_headers.get(field)
        if not h:
            return ''
        v = row_dict.get(h, '')
        return '' if str(v).strip().lower() in BLANK_TOKENS else str(v).strip()
    return get


def open_row_stream(file_path, ext):
    """Returns (headers: list[str], row_iterator: generator yielding dict per data row)."""
    if ext == 'csv':
        f = open(file_path, 'r', encoding='utf-8-sig', errors='replace', newline='')
        reader = csv.reader(f)
        try:
            first = next(reader)
        except StopIteration:
            f.close()
            return [], iter(())
        headers = [str(c).strip() for c in first]

        def gen():
            try:
                for raw_row in reader:
                    row_dict = {}
                    for i, h in enumerate(headers):
                        val = raw_row[i] if i < len(raw_row) else ''
                        row_dict[h] = str(val).strip()
                    yield row_dict
            finally:
                f.close()
        return headers, gen()

    if ext == 'xlsx':
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        row_iter = ws.iter_rows(values_only=True)
        try:
            first = next(row_iter)
        except StopIteration:
            wb.close()
            return [], iter(())
        headers = [str(c).strip() if c is not None else '' for c in first]

        def gen():
            try:
                for raw_row in row_iter:
                    row_dict = {}
                    for i, h in enumerate(headers):
                        val = raw_row[i] if i < len(raw_row) else None
                        row_dict[h] = '' if val is None else str(val).strip()
                    yield row_dict
            finally:
                wb.close()
        return headers, gen()

    # Legacy .xls: whole-sheet fallback (the format itself caps out around ~65k rows)
    import pandas as pd
    df = pd.read_excel(file_path, dtype=str).fillna('')
    headers = [str(c).strip() for c in df.columns]

    def gen():
        for _, row in df.iterrows():
            row_dict = {}
            for c, h in zip(df.columns, headers):
                v = str(row[c]).strip()
                row_dict[h] = '' if v.lower() in ('nan', 'none', 'null') else v
            yield row_dict
    return headers, gen()


def quick_count_total_rows(file_path, ext):
    if ext == 'csv':
        with open(file_path, 'rb') as f:
            count = sum(1 for _ in f)
        return max(0, count - 1)
    if ext == 'xlsx':
        wb = load_workbook(file_path, read_only=True)
        ws = wb.worksheets[0]
        total = max(0, (ws.max_row or 1) - 1)
        wb.close()
        return total
    import pandas as pd
    df = pd.read_excel(file_path, dtype=str)
    return len(df)


def _is_blank_row(row_dict):
    for v in row_dict.values():
        if str(v).strip().lower() not in BLANK_TOKENS:
            return False
    return True


def _validate(name, sku, cost_price, selling_price, stock):
    errors = []
    if not name and not sku:
        errors.append("Missing Product Name or Code")
    if cost_price < 0:
        errors.append("Cost Price cannot be negative")
    if selling_price < 0:
        errors.append("Selling Price cannot be negative")
    if stock < 0:
        errors.append("Stock quantity cannot be negative")
    return errors


def _process_chunk(batch, chunk_rows, resolved, duplicate_strategy,
                    existing_skus, existing_barcodes, seen_barcodes_in_file):
    """Validates + writes one chunk. Returns (imported, failed, duplicate) deltas."""
    imported = failed = duplicate = 0
    new_products = []
    update_skus = []
    replace_skus = []
    error_logs = []
    seen_skus_in_chunk = set()

    for row_no, row_dict in chunk_rows:
        if _is_blank_row(row_dict):
            continue

        get = field_getter(resolved, row_dict)
        sku = get('sku') or get('model_no')
        name = get('name') or get('name_local')
        barcode = get('barcode')
        category = get('category') or 'Frames'
        brand = get('brand') or 'Generic'
        supplier = get('supplier')
        cost_price = safe_float(get('cost_price'), 0.0)
        selling_price = safe_float(get('selling_price') or get('cost_price'), 0.0)
        stock = safe_int(get('stock'), 1)

        if not sku and not name:
            continue

        if not sku:
            sku = f"SKU-{uuid.uuid4().hex[:8].upper()}"
        sku = sku[:140]

        errors = _validate(name, sku, cost_price, selling_price, stock)
        if errors:
            failed += 1
            error_logs.append(ImportErrorLog(
                batch=batch, sheet_name='Sheet 1', row_number=row_no,
                product_code=sku, product_name=name, field_name='Validation',
                error_type='VALIDATION_FAILED', error_message=", ".join(errors), raw_data=row_dict
            ))
            continue

        is_dup = sku in existing_skus or sku in seen_skus_in_chunk
        clean_barcode = barcode[:140] if barcode else None
        if clean_barcode:
            if clean_barcode in existing_barcodes or clean_barcode in seen_barcodes_in_file:
                clean_barcode = None
            else:
                seen_barcodes_in_file.add(clean_barcode)
                existing_barcodes.add(clean_barcode)

        row_payload = dict(
            name=(name or sku)[:250],
            category_name=(infer_category_name(category, name) or category)[:140],
            brand_name=brand[:140],
            supplier_name=supplier[:140] if supplier else '',
            cost_price=cost_price,
            retail_price=selling_price,
            stock=stock,
            extra_data=row_dict,
        )

        if is_dup:
            duplicate += 1
            if duplicate_strategy == 'SKIP':
                continue
            elif duplicate_strategy == 'UPDATE':
                update_skus.append((sku, clean_barcode, row_payload))
            elif duplicate_strategy == 'REPLACE':
                replace_skus.append(sku)
                new_products.append(Product(
                    sku=sku, barcode=clean_barcode, opening_stock=stock,
                    import_batch=batch, is_active=True, **row_payload
                ))
                imported += 1
            elif duplicate_strategy == 'DUPLICATE':
                dup_sku = f"{sku}-DUP{uuid.uuid4().hex[:4].upper()}"[:140]
                dup_barcode = f"{clean_barcode}-DUP"[:140] if clean_barcode else None
                row_payload['name'] = f"{row_payload['name']} (Duplicate)"[:250]
                new_products.append(Product(
                    sku=dup_sku, barcode=dup_barcode, opening_stock=stock,
                    import_batch=batch, is_active=True, **row_payload
                ))
                imported += 1
        else:
            seen_skus_in_chunk.add(sku)
            existing_skus.add(sku)
            new_products.append(Product(
                sku=sku, barcode=clean_barcode, opening_stock=stock,
                import_batch=batch, is_active=True, **row_payload
            ))
            imported += 1

    with transaction.atomic():
        if replace_skus:
            Product.objects.filter(sku__in=replace_skus).delete()
        if new_products:
            Product.objects.bulk_create(new_products, batch_size=CHUNK_SIZE)
        if update_skus:
            existing_objs = {p.sku: p for p in Product.objects.filter(sku__in=[s for s, _, _ in update_skus])}
            to_update = []
            for sku, clean_barcode, payload in update_skus:
                p = existing_objs.get(sku)
                if not p:
                    continue
                for field, value in payload.items():
                    setattr(p, field, value)
                if clean_barcode:
                    p.barcode = clean_barcode
                p.import_batch = batch
                to_update.append(p)
            if to_update:
                Product.objects.bulk_update(
                    to_update,
                    fields=['name', 'category_name', 'brand_name', 'supplier_name',
                            'cost_price', 'retail_price', 'stock', 'extra_data',
                            'barcode', 'import_batch'],
                    batch_size=CHUNK_SIZE
                )
            imported += len(to_update)
        if error_logs:
            ImportErrorLog.objects.bulk_create(error_logs, batch_size=CHUNK_SIZE)

    return imported, failed, duplicate


def run_import_job(batch_id):
    close_old_connections()
    try:
        batch = ImportBatch.objects.get(id=batch_id)
    except ImportBatch.DoesNotExist:
        return

    start_time = time.time()
    try:
        batch.status = 'PROCESSING'
        batch.save(update_fields=['status'])

        ext = detect_ext(batch.file_name)
        file_path = batch.file.path

        total_rows = quick_count_total_rows(file_path, ext)
        headers, row_gen = open_row_stream(file_path, ext)
        resolved = resolve_field_headers(headers)

        batch.total_rows = total_rows
        batch.column_headers = headers
        batch.save(update_fields=['total_rows', 'column_headers'])

        existing_skus = set(Product.objects.exclude(sku__isnull=True).values_list('sku', flat=True))
        existing_barcodes = set(
            Product.objects.exclude(barcode__isnull=True).exclude(barcode='').values_list('barcode', flat=True)
        )
        seen_barcodes_in_file = set()

        processed = imported = failed = duplicate = 0
        chunk = []

        def flush(chunk_rows):
            nonlocal processed, imported, failed, duplicate
            if not chunk_rows:
                return
            try:
                d_imported, d_failed, d_duplicate = _process_chunk(
                    batch, chunk_rows, resolved, batch.duplicate_strategy,
                    existing_skus, existing_barcodes, seen_barcodes_in_file
                )
            except Exception as chunk_err:
                ImportErrorLog.objects.bulk_create([
                    ImportErrorLog(
                        batch=batch, sheet_name='Sheet 1', row_number=r_no,
                        product_code='', error_type='CHUNK_FAILED',
                        error_message=str(chunk_err), raw_data=r_data
                    ) for r_no, r_data in chunk_rows
                ], batch_size=CHUNK_SIZE)
                d_imported, d_failed, d_duplicate = 0, len(chunk_rows), 0

            processed += len(chunk_rows)
            imported += d_imported
            failed += d_failed
            duplicate += d_duplicate
            batch.processed_rows = processed
            batch.imported_rows = imported
            batch.failed_rows = failed
            batch.duplicate_rows = duplicate
            batch.save(update_fields=['processed_rows', 'imported_rows', 'failed_rows', 'duplicate_rows'])

        row_no = 1  # header was row 1
        for row_dict in row_gen:
            row_no += 1
            chunk.append((row_no, row_dict))
            if len(chunk) >= CHUNK_SIZE:
                flush(chunk)
                chunk = []
        flush(chunk)

        processing_time = round(time.time() - start_time, 2)
        batch.processing_time = processing_time
        batch.status = 'SUCCESS' if failed == 0 else ('PARTIAL' if imported > 0 else 'FAILED')
        batch.remarks = (
            f"Imported {imported}, skipped/duplicate {duplicate}, failed {failed} "
            f"of {processed} rows read in {processing_time}s."
        )
        batch.logs = (batch.logs or []) + [
            f"{datetime.datetime.now().strftime('%H:%M:%S')} - Import finished: "
            f"{imported} imported, {duplicate} duplicates, {failed} failed, {processing_time}s"
        ]
        batch.save(update_fields=['processing_time', 'status', 'remarks', 'logs'])
    except Exception as e:
        try:
            batch.status = 'FAILED'
            batch.remarks = f"Import crashed: {e}"
            batch.save(update_fields=['status', 'remarks'])
        except Exception:
            pass
    finally:
        connections.close_all()
